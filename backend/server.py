from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Header
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import logging
import bcrypt
import jwt
import secrets
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from enum import Enum
import io
import csv
import asyncio
from openpyxl import load_workbook
from email_service import send_enrollment_welcome_email

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL')
if not mongo_url:
    logger.error("MONGO_URL environment variable is NOT set!")
    # We allow it to continue so we can actually see the log in Render
    db = None
else:
    client = AsyncIOMotorClient(mongo_url)
    db_name = os.environ.get('DB_NAME', 'academic_crm')
    db = client[db_name]
    logger.info(f"Connected to MongoDB: {db_name}")

# JWT Config
JWT_ALGORITHM = "HS256"
JWT_SECRET = os.environ.get("JWT_SECRET", secrets.token_hex(32))

# Create the main app
app = FastAPI(title="Raffles University Academic CRM")
api_router = APIRouter()

# ==================== ENUMS ====================
class UserRole(str, Enum):
    ADMIN = "admin"
    DEAN = "dean"
    DEAN_ACADEMICS = "dean_academics"
    HOD = "hod"
    FACULTY = "faculty"
    STAFF = "staff"
    STUDENT = "student"

class AttendanceStatus(str, Enum):
    PRESENT = "present"
    ABSENT = "absent"
    LATE = "late"

class SubjectType(str, Enum):
    CORE = "core"
    ELECTIVE = "elective"
    LAB = "lab"
    PRACTICAL = "practical"

# ==================== PYDANTIC MODELS ====================
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: UserRole = UserRole.STUDENT
    department_id: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    department_id: Optional[str] = None
    created_at: datetime

class DepartmentCreate(BaseModel):
    name: str
    code: str
    head_id: Optional[str] = None

class DepartmentResponse(BaseModel):
    id: str
    name: str
    code: str
    head_id: Optional[str] = None
    created_at: datetime

class ProgramCreate(BaseModel):
    name: str
    code: str
    department_id: str
    duration_years: int
    total_semesters: int

class ProgramResponse(BaseModel):
    id: str
    name: str
    code: str
    department_id: str
    department_name: Optional[str] = None
    duration_years: int
    total_semesters: int
    created_at: datetime

class AdmissionStudentItem(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None
    enrollmentNo: str
    category: Optional[str] = None

class AdmissionSyncRequest(BaseModel):
    source: str
    session: str
    program: str
    department: str
    students: List[AdmissionStudentItem]

class IntegrationLog(BaseModel):
    source: str
    timestamp: datetime
    totalReceived: int
    inserted: int
    duplicates: int
    session: str
    program: str
    status: str

class StudentCreate(BaseModel):
    name: str
    enrollment_number: str
    email: EmailStr
    mobile_number: Optional[str] = None
    program_id: str
    academic_session: str
    category: Optional[str] = None # General, OBC, SC, ST etc.
    semester: int = 1

class StudentResponse(BaseModel):
    id: str
    student_id: str
    name: str
    enrollment_number: str
    email: str
    mobile_number: Optional[str] = None
    program_id: str
    program_name: Optional[str] = None
    department_id: Optional[str] = None
    department_name: Optional[str] = None
    academic_session: str
    category: Optional[str] = None
    semester: int
    user_id: Optional[str] = None
    created_at: datetime

class SubjectCreate(BaseModel):
    name: str
    code: str
    type: str # Core, Elective, Lab, Practical
    credits: float
    program_id: str
    semester: int

class SubjectResponse(BaseModel):
    id: str
    name: str
    code: str
    type: str
    credits: float
    program_id: str
    program_name: Optional[str] = None
    semester: int
    created_at: datetime

class ElectiveSelection(BaseModel):
    student_id: str
    subject_id: str
    semester: int
    status: str = "pending" # pending, approved, rejected

class SemesterRecordResponse(BaseModel):
    id: str
    student_id: str
    program_id: str
    semester: int
    session: str
    subjects: List[SubjectResponse]
    status: str = "ACTIVE"
    created_at: datetime

class ProgressionRequest(BaseModel):
    target_semester: int

# ==================== HELPER FUNCTIONS ====================
def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode("utf-8"), salt).decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=24),
        "type": "access"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["id"] = str(user["_id"])
        user.pop("_id", None)
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_roles(allowed_roles: List[str]):
    async def role_checker(request: Request):
        user = await get_current_user(request)
        if user["role"] not in allowed_roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return role_checker

def generate_student_id(program_code: str, session: str, count: int) -> str:
    year = session.split("-")[0][-2:]
    return f"RU{year}{program_code}{str(count).zfill(4)}"

# ==================== PROGRAMS & STUDENT HISTORY ====================
async def get_student_history(student_id: str, user: dict = Depends(get_current_user)):
    # Data isolation
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
        
    if user["role"] == "student" and user["email"].lower() != student["email"].lower():
        raise HTTPException(status_code=403, detail="Unauthorized to view this history")
    
    # We fetch records from a new 'semester_history' collection
    history = await db.semester_history.find({"student_id": student_id}).sort("semester", 1).to_list(100)
    
    result = []
    for h in history:
        # Resolve subjects for this record
        subjects = []
        for s_id in h.get("subject_ids", []):
            s = await db.subjects.find_one({"_id": ObjectId(s_id)})
            if s:
                prog = await db.programs.find_one({"_id": ObjectId(s["program_id"])})
                subjects.append(SubjectResponse(
                    id=str(s["_id"]),
                    name=s["name"],
                    code=s["code"],
                    type=s.get("type", "Core"),
                    credits=s.get("credits", 0),
                    program_id=s["program_id"],
                    program_name=prog["name"] if prog else "Unknown",
                    semester=s["semester"],
                    created_at=s["created_at"]
                ))
        
        result.append(SemesterRecordResponse(
            id=str(h["_id"]),
            student_id=student_id,
            program_id=h["program_id"],
            semester=h["semester"],
            session=h["session"],
            subjects=subjects,
            status=h.get("status", "COMPLETED"),
            created_at=h["created_at"]
        ))
    return result

@api_router.post("/students/{student_id}/progress")
async def progress_student(student_id: str, data: ProgressionRequest, user: dict = Depends(require_roles(["admin", "dean", "dean_academics"]))):
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    current_sem = student["semester"]
    target_sem = data.target_semester
    
    # 1. Create a Snapshot of Current Semester before progressing
    current_subjects = await db.subjects.find({"program_id": student["program_id"], "semester": current_sem}).to_list(100)
    
    snapshot = {
        "student_id": student_id,
        "program_id": student["program_id"],
        "semester": current_sem,
        "session": student["academic_session"],
        "subject_ids": [str(s["_id"]) for s in current_subjects],
        "status": "COMPLETED",
        "created_at": datetime.now(timezone.utc)
    }
    await db.semester_history.insert_one(snapshot)
    
    # 2. Update Student to Target Semester
    await db.students.update_one(
        {"_id": ObjectId(student_id)},
        {"$set": {"semester": target_sem}}
    )
    
    return {"message": f"Student progressed from Semester {current_sem} to {target_sem}", "snapshot_created": True}

class TimetableSlotCreate(BaseModel):
    day: str
    start_time: str
    end_time: str
    subject_id: str
    faculty_id: str
    room: str
    program_id: str
    semester: int

class TimetableSlotResponse(BaseModel):
    id: str
    day: str
    start_time: str
    end_time: str
    subject_id: str
    subject_name: Optional[str] = None
    faculty_id: str
    faculty_name: Optional[str] = None
    room: str
    program_id: str
    program_name: Optional[str] = None
    semester: int
    created_at: datetime

class AttendanceCreate(BaseModel):
    student_id: str
    subject_id: str
    date: str
    status: AttendanceStatus

class AttendanceBulkCreate(BaseModel):
    subject_id: str
    date: str
    records: List[Dict[str, str]]

class AttendanceResponse(BaseModel):
    id: str
    student_id: str
    student_name: Optional[str] = None
    subject_id: str
    subject_name: Optional[str] = None
    date: str
    status: str
    marked_by: Optional[str] = None
    created_at: datetime

class ElectiveCreate(BaseModel):
    name: str
    code: str
    credits: int
    program_id: str
    semester: int
    max_students: int = 60

class ElectiveResponse(BaseModel):
    id: str
    name: str
    code: str
    credits: int
    program_id: str
    program_name: Optional[str] = None
    semester: int
    max_students: int
    created_at: datetime

class BatchCreate(BaseModel):
    elective_id: str
    faculty_id: str
    batch_name: str

class BatchResponse(BaseModel):
    id: str
    elective_id: str
    elective_name: Optional[str] = None
    faculty_id: str
    faculty_name: Optional[str] = None
    batch_name: str
    student_ids: List[str] = []
    created_at: datetime

# ==================== AUTH ENDPOINTS ====================
@api_router.post("/auth/register")
async def register(user_data: UserCreate, response: Response):
    email = user_data.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_doc = {
        "email": email,
        "password_hash": hash_password(user_data.password),
        "name": user_data.name,
        "role": user_data.role.value,
        "department_id": user_data.department_id,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)
    
    access_token = create_access_token(user_id, email, user_data.role.value)
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    
    return {"id": user_id, "email": email, "name": user_data.name, "role": user_data.role.value}

@api_router.post("/auth/login")
async def login(user_data: UserLogin, response: Response):
    email = user_data.email.lower()
    user = await db.users.find_one({"email": email})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(user_data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email, user["role"])
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    
    return {"id": user_id, "email": user["email"], "name": user["name"], "role": user["role"]}

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out successfully"}

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    return user

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    """Issue a new access_token using a valid refresh_token cookie."""
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user_id = payload["sub"]
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        new_access = create_access_token(user_id, user["email"], user["role"])
        response.set_cookie(
            key="access_token", value=new_access,
            httponly=True, secure=False, samesite="lax",
            max_age=86400, path="/"
        )
        return {"message": "Token refreshed"}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expired — please log in again")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ==================== DEPARTMENT ENDPOINTS ====================
@api_router.post("/departments", response_model=DepartmentResponse)
async def create_department(dept: DepartmentCreate, user: dict = Depends(require_roles(["admin", "dean"]))):
    existing = await db.departments.find_one({"code": dept.code})
    if existing:
        raise HTTPException(status_code=400, detail="Department code already exists")
    
    doc = {
        "name": dept.name,
        "code": dept.code,
        "head_id": dept.head_id,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.departments.insert_one(doc)
    return DepartmentResponse(id=str(result.inserted_id), **{k:v for k,v in doc.items() if k != "_id"})

@api_router.get("/departments", response_model=List[DepartmentResponse])
async def get_departments(user: dict = Depends(get_current_user)):
    depts = await db.departments.find({}).to_list(100)
    return [DepartmentResponse(id=str(d["_id"]), name=d["name"], code=d["code"], head_id=d.get("head_id"), created_at=d["created_at"]) for d in depts]

@api_router.get("/departments/{dept_id}", response_model=DepartmentResponse)
async def get_department(dept_id: str, user: dict = Depends(get_current_user)):
    dept = await db.departments.find_one({"_id": ObjectId(dept_id)})
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    return DepartmentResponse(id=str(dept["_id"]), name=dept["name"], code=dept["code"], head_id=dept.get("head_id"), created_at=dept["created_at"])

@api_router.delete("/departments/{dept_id}")
async def delete_department(dept_id: str, user: dict = Depends(require_roles(["admin"]))):
    result = await db.departments.delete_one({"_id": ObjectId(dept_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Department not found")
    return {"message": "Department deleted"}

# ==================== PROGRAMS & STUDENT HISTORY ====================
# Global roles that can see data across all departments
GLOBAL_ROLES = ["admin", "dean_academics"]

@api_router.get("/students", response_model=List[StudentResponse])
async def get_students(
    program_id: Optional[str] = None,
    academic_session: Optional[str] = None,
    search: Optional[str] = None,
    user: dict = Depends(require_roles(["admin", "dean", "dean_academics", "hod", "faculty", "staff", "student"]))
):
    query = {}
    
    # Apply Department/Role Filtering
    if user["role"] == "student":
        # Students can only see their own record
        query["email"] = user["email"].lower()
    elif user["role"] not in GLOBAL_ROLES:
        # Deans, HODs, Faculty, Staff can only see their department
        if user.get("department_id"):
            query["department_id"] = user["department_id"]
    
    if program_id:
        query["program_id"] = program_id
    if academic_session:
        query["academic_session"] = academic_session
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"student_id": {"$regex": search, "$options": "i"}},
            {"enrollment_number": {"$regex": search, "$options": "i"}}
        ]
        
    students = await db.students.find(query).sort("created_at", -1).to_list(1000)

    # ── Batch-resolve program & department names (eliminates N+1) ──────────
    program_ids = {ObjectId(s["program_id"]) for s in students if s.get("program_id") and ObjectId.is_valid(s["program_id"])}
    dept_ids    = {ObjectId(s["department_id"]) for s in students if s.get("department_id") and ObjectId.is_valid(s["department_id"])}

    prog_map = {str(p["_id"]): p async for p in db.programs.find({"_id": {"$in": list(program_ids)}})}
    dept_map = {str(d["_id"]): d async for d in db.departments.find({"_id": {"$in": list(dept_ids)}})}

    result = []
    for s in students:
        prog = prog_map.get(s.get("program_id", ""))
        dept = dept_map.get(s.get("department_id", ""))
        result.append(StudentResponse(
            id=str(s["_id"]),
            student_id=s["student_id"],
            name=s["name"],
            enrollment_number=s["enrollment_number"],
            email=s["email"],
            mobile_number=s.get("mobile_number", s.get("phone")),
            program_id=s["program_id"],
            program_name=prog["name"] if prog else "Unknown",
            department_id=s.get("department_id"),
            department_name=dept["name"] if dept else "Unknown",
            academic_session=s["academic_session"],
            semester=s["semester"],
            category=s.get("category"),
            created_at=s["created_at"]
        ))
    return result

@api_router.get("/programs", response_model=List[ProgramResponse])
async def get_programs(
    department_id: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    
    # Restrict Deans/HODs to their own department's programs
    if user["role"] not in GLOBAL_ROLES and user["role"] != "student":
        if user.get("department_id"):
            query["department_id"] = user["department_id"]
            
    if department_id:
        query["department_id"] = department_id
        
    programs = await db.programs.find(query).to_list(200)
    
    result = []
    for p in programs:
        dept = await db.departments.find_one({"_id": ObjectId(p["department_id"]) if ObjectId.is_valid(p["department_id"]) else p["department_id"]})
        result.append(ProgramResponse(
            id=str(p["_id"]),
            name=p["name"],
            code=p["code"],
            department_id=p["department_id"],
            department_name=dept["name"] if dept else "Unknown",
            duration_years=p.get("duration_years", 3),
            total_semesters=p.get("total_semesters", 6),
            created_at=p["created_at"]
        ))
    return result

@api_router.post("/programs", response_model=ProgramResponse)
async def create_program(prog: ProgramCreate, user: dict = Depends(require_roles(["admin", "dean", "dean_academics"]))):
    existing = await db.programs.find_one({"code": prog.code})
    if existing:
        raise HTTPException(status_code=400, detail="Program code already exists")
    
    dept = await db.departments.find_one({"_id": ObjectId(prog.department_id)})
    if not dept:
        raise HTTPException(status_code=400, detail="Department not found")
    
    doc = {
        "name": prog.name,
        "code": prog.code,
        "department_id": prog.department_id,
        "duration_years": prog.duration_years,
        "total_semesters": prog.total_semesters,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.programs.insert_one(doc)
    return ProgramResponse(id=str(result.inserted_id), department_name=dept["name"], **{k:v for k,v in doc.items()})

@api_router.get("/programs", response_model=List[ProgramResponse])
async def get_programs(department_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    query = {}
    if department_id:
        query["department_id"] = department_id
    
    programs = await db.programs.find(query).to_list(100)
    result = []
    for p in programs:
        dept = await db.departments.find_one({"_id": ObjectId(p["department_id"])})
        result.append(ProgramResponse(
            id=str(p["_id"]),
            name=p["name"],
            code=p["code"],
            department_id=p["department_id"],
            department_name=dept["name"] if dept else None,
            duration_years=p["duration_years"],
            total_semesters=p["total_semesters"],
            created_at=p["created_at"]
        ))
    return result

@api_router.get("/programs/{prog_id}", response_model=ProgramResponse)
async def get_program(prog_id: str, user: dict = Depends(get_current_user)):
    prog = await db.programs.find_one({"_id": ObjectId(prog_id)})
    if not prog:
        raise HTTPException(status_code=404, detail="Program not found")
    dept = await db.departments.find_one({"_id": ObjectId(prog["department_id"])})
    return ProgramResponse(
        id=str(prog["_id"]),
        name=prog["name"],
        code=prog["code"],
        department_id=prog["department_id"],
        department_name=dept["name"] if dept else None,
        duration_years=prog["duration_years"],
        total_semesters=prog["total_semesters"],
        created_at=prog["created_at"]
    )

@api_router.delete("/programs/{prog_id}")
async def delete_program(prog_id: str, user: dict = Depends(require_roles(["admin", "dean"]))):
    result = await db.programs.delete_one({"_id": ObjectId(prog_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Program not found")
    return {"message": "Program deleted"}

# ==================== STUDENT ENDPOINTS ====================
@api_router.post("/students", response_model=StudentResponse)
async def create_student(student: StudentCreate, user: dict = Depends(require_roles(["admin", "dean", "staff"]))):
    existing = await db.students.find_one({"enrollment_number": student.enrollment_number})
    if existing:
        raise HTTPException(status_code=400, detail="Enrollment number already exists")
    
    prog = await db.programs.find_one({"_id": ObjectId(student.program_id)})
    if not prog:
        raise HTTPException(status_code=400, detail="Program not found")
    
    dept = await db.departments.find_one({"_id": ObjectId(prog["department_id"])})
    
    # Generate unique student ID
    count = await db.students.count_documents({"program_id": student.program_id, "academic_session": student.academic_session})
    student_id = generate_student_id(prog["code"], student.academic_session, count + 1)
    
    doc = {
        "student_id": student_id,
        "name": student.name,
        "enrollment_number": student.enrollment_number,
        "email": student.email.lower(),
        "mobile_number": student.mobile_number,
        "program_id": student.program_id,
        "department_id": prog["department_id"],
        "academic_session": student.academic_session,
        "category": student.category,
        "semester": student.semester,
        "user_id": None,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.students.insert_one(doc)

    # ── Fire welcome email (non-blocking, errors are logged not raised) ──
    asyncio.create_task(send_enrollment_welcome_email(
        to_email=student.email,
        student_name=student.name,
        student_id=student_id,
        program_name=prog["name"],
        semester=student.semester,
        enrollment_number=student.enrollment_number,
        academic_session=student.academic_session,
        department_name=dept["name"] if dept else "",
    ))

    return StudentResponse(
        id=str(result.inserted_id),
        student_id=student_id,
        name=student.name,
        enrollment_number=student.enrollment_number,
        email=student.email,
        mobile_number=student.mobile_number,
        program_id=student.program_id,
        program_name=prog["name"],
        department_id=prog["department_id"],
        department_name=dept["name"] if dept else None,
        academic_session=student.academic_session,
        category=student.category,
        semester=student.semester,
        created_at=doc["created_at"]
    )

@api_router.post("/students/bulk-import")
async def bulk_import_students(file: UploadFile = File(...), user: dict = Depends(require_roles(["admin", "dean", "staff"]))):
    content = await file.read()
    filename = file.filename.lower() if file.filename else ""
    
    rows = []
    
    # Parse based on file type
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        # Excel file
        try:
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = None
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [str(h).strip().lower() if h else '' for h in row]
                else:
                    if headers and any(cell is not None for cell in row):
                        row_dict = {}
                        for i, val in enumerate(row):
                            if i < len(headers) and headers[i]:
                                row_dict[headers[i]] = str(val).strip() if val is not None else ''
                        rows.append(row_dict)
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")
    else:
        # CSV file
        try:
            decoded = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded))
            for row in reader:
                rows.append({k.lower().strip(): v for k, v in row.items()})
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading CSV file: {str(e)}")
    
    imported = 0
    errors = []
    
    # 1. Pre-fetch all programs
    all_progs = await db.programs.find().to_list(1000)
    prog_map = {p["code"].upper(): p for p in all_progs}
    
    # 2. Extract all incoming enrollments to check duplicates efficiently
    incoming_enrollments = []
    for row in rows:
        enh = str(row.get("enrollment_number", row.get("enrollment", ""))).strip()
        if enh: incoming_enrollments.append(enh)
    
    # Pre-fetch existing students to easily do duplicate checks
    existing_students = await db.students.find({"enrollment_number": {"$in": incoming_enrollments}}).to_list(None)
    existing_enrollments = {s["enrollment_number"] for s in existing_students}
    
    # 3. Maintain a local counter for student IDs
    session_prog_counts = {}
    docs_to_insert = []
    
    for row_idx, row in enumerate(rows):
        try:
            program_code = str(row.get("program_code", row.get("program", ""))).strip().upper()
            prog = prog_map.get(program_code)
            if not prog:
                errors.append(f"Row {row_idx + 2}: Program not found - {program_code}")
                continue
            
            enrollment = str(row.get("enrollment_number", row.get("enrollment", ""))).strip()
            if not enrollment:
                errors.append(f"Row {row_idx + 2}: Missing enrollment number")
                continue
                
            if enrollment in existing_enrollments:
                errors.append(f"Row {row_idx + 2}: Enrollment number already exists - {enrollment}")
                continue
            
            # Add to local map to prevent duplicates inside the uploaded file itself
            existing_enrollments.add(enrollment)
            
            session = str(row.get("academic_session", row.get("session", "2025-2029"))).strip()
            
            prog_id_str = str(prog["_id"])
            count_key = f"{prog_id_str}_{session}"
            if count_key not in session_prog_counts:
                # Only hit the DB once per program-session combination
                count_docs = await db.students.count_documents({"program_id": prog_id_str, "academic_session": session})
                session_prog_counts[count_key] = count_docs
                
            session_prog_counts[count_key] += 1
            student_id = generate_student_id(prog["code"], session, session_prog_counts[count_key])
            
            name = str(row.get("name", row.get("student_name", ""))).strip()
            email = str(row.get("email", row.get("student_email", ""))).strip().lower()
            
            if not name:
                errors.append(f"Row {row_idx + 2}: Missing student name")
                continue
            
            doc = {
                "student_id": student_id,
                "name": name,
                "enrollment_number": enrollment,
                "email": email or f"{enrollment.lower()}@raffles.edu.in",
                "mobile_number": str(row.get("mobile_number", row.get("mobile", row.get("phone", "")))).strip() or None,
                "program_id": prog_id_str,
                "department_id": prog["department_id"],
                "academic_session": session,
                "category": str(row.get("category", "")).strip() or None,
                "semester": int(row.get("semester", 1) or 1),
                "user_id": None,
                "created_at": datetime.now(timezone.utc)
            }
            docs_to_insert.append(doc)
            imported += 1
        except Exception as e:
            errors.append(f"Row {row_idx + 2}: {str(e)}")
            
    if docs_to_insert:
        await db.students.insert_many(docs_to_insert)
    
    return {"imported": imported, "errors": errors, "total_rows": len(rows)}

@api_router.post("/integration/admissions/sync")
async def admission_sync(data: AdmissionSyncRequest, x_api_key: str = Header(None)):
    expected_api_key = os.environ.get("ADMISSION_CRM_API_KEY")
    
    if not expected_api_key or x_api_key != expected_api_key:
        raise HTTPException(status_code=401, detail="Invalid API key or source")
    
    if data.source != "admission_crm":
        raise HTTPException(status_code=400, detail="Unauthorized source")
    
    # 1. Resolve Program & Department
    prog = await db.programs.find_one({
        "$or": [{"code": data.program.upper()}, {"name": data.program}]
    })
    if not prog:
        raise HTTPException(status_code=400, detail=f"Program not found: {data.program}")
    
    # Session Validation
    try:
        start_year, end_year = map(int, data.session.split('-'))
        if end_year - start_year != prog["duration_years"]:
            raise HTTPException(
                status_code=400, 
                detail=f"Session duration mismatch. The program '{prog['name']}' requires a {prog['duration_years']}-year session (e.g., {start_year}-{start_year + prog['duration_years']})."
            )
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="Invalid session format. Expected 'YYYY-YYYY' (e.g., 2025-2029).")
    
    inserted = 0
    duplicates = 0
    total_received = len(data.students)
    
    for student in data.students:
        # Check for duplicate by enrollment number
        existing = await db.students.find_one({"enrollment_number": student.enrollmentNo})
        if existing:
            duplicates += 1
            continue
            
        # Transform & Map
        student_count = await db.students.count_documents({"program_id": str(prog["_id"]), "academic_session": data.session})
        student_id = generate_student_id(prog["code"], data.session, student_count + 1)
        
        doc = {
            "student_id": student_id,
            "name": student.name,
            "enrollment_number": student.enrollmentNo,
            "email": student.email.lower(),
            "phone": student.phone,
            "program_id": str(prog["_id"]),
            "department_id": prog["department_id"],
            "academic_session": data.session,
            "category": student.category,
            "semester": 1, # Default to first semester
            "academicStatus": "ACTIVE",
            "user_id": None,
            "created_at": datetime.now(timezone.utc)
        }
        
        await db.students.insert_one(doc)
        inserted += 1
        
    # Transaction Logging (Audit Trail)
    log_doc = {
        "source": data.source,
        "timestamp": datetime.now(timezone.utc),
        "totalReceived": total_received,
        "inserted": inserted,
        "duplicates": duplicates,
        "session": data.session,
        "program": data.program,
        "status": "success"
    }
    await db.integration_logs.insert_one(log_doc)
    
    return {
        "status": "success",
        "message": "Students synced successfully",
        "totalReceived": total_received,
        "inserted": inserted,
        "duplicates": duplicates
    }

@api_router.get("/integration/logs", response_model=List[IntegrationLog])
async def get_integration_logs(user: dict = Depends(require_roles(["admin", "dean_academics"]))):
    logs = await db.integration_logs.find().sort("timestamp", -1).to_list(100)
    return logs

@api_router.get("/students", response_model=List[StudentResponse])
async def get_students(
    program_id: Optional[str] = None,
    department_id: Optional[str] = None,
    academic_session: Optional[str] = None,
    semester: Optional[int] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if program_id:
        query["program_id"] = program_id
    if department_id:
        query["department_id"] = department_id
    if academic_session:
        query["academic_session"] = academic_session
    if semester:
        query["semester"] = semester
    
    students = await db.students.find(query).to_list(1000)
    result = []
    
    for s in students:
        prog = await db.programs.find_one({"_id": ObjectId(s["program_id"])}) if s.get("program_id") else None
        dept = await db.departments.find_one({"_id": ObjectId(s["department_id"])}) if s.get("department_id") else None
        result.append(StudentResponse(
            id=str(s["_id"]),
            student_id=s["student_id"],
            name=s["name"],
            enrollment_number=s["enrollment_number"],
            email=s["email"],
            program_id=s["program_id"],
            program_name=prog["name"] if prog else None,
            department_id=s.get("department_id"),
            department_name=dept["name"] if dept else None,
            academic_session=s["academic_session"],
            category=s.get("category"),
            semester=s["semester"],
            user_id=s.get("user_id"),
            created_at=s["created_at"]
        ))
    return result

@api_router.get("/students/{student_id}", response_model=StudentResponse)
async def get_student(student_id: str, user: dict = Depends(get_current_user)):
    s = await db.students.find_one({"_id": ObjectId(student_id)})
    if not s:
        raise HTTPException(status_code=404, detail="Student not found")
    prog = await db.programs.find_one({"_id": ObjectId(s["program_id"])}) if s.get("program_id") else None
    dept = await db.departments.find_one({"_id": ObjectId(s["department_id"])}) if s.get("department_id") else None
    return StudentResponse(
        id=str(s["_id"]),
        student_id=s["student_id"],
        name=s["name"],
        enrollment_number=s["enrollment_number"],
        email=s["email"],
        program_id=s["program_id"],
        program_name=prog["name"] if prog else None,
        department_id=s.get("department_id"),
        department_name=dept["name"] if dept else None,
        academic_session=s["academic_session"],
        category=s.get("category"),
        semester=s["semester"],
        user_id=s.get("user_id"),
        created_at=s["created_at"]
    )

@api_router.delete("/students/{student_id}")
async def delete_student(student_id: str, user: dict = Depends(require_roles(["admin", "dean"]))):
    result = await db.students.delete_one({"_id": ObjectId(student_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Student not found")
    return {"message": "Student deleted"}

# ==================== SUBJECT ENDPOINTS ====================
@api_router.post("/subjects", response_model=SubjectResponse)
async def create_subject(subj: SubjectCreate, user: dict = Depends(require_roles(["admin", "dean", "dean_academics"]))):
    # Check for duplicate code
    existing = await db.subjects.find_one({"code": subj.code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Subject code '{subj.code}' already exists")

    prog = await db.programs.find_one({"_id": ObjectId(subj.program_id) if ObjectId.is_valid(subj.program_id) else subj.program_id})
    if not prog:
        raise HTTPException(status_code=404, detail="Program not found")

    # Department isolation for Deans
    if user["role"] not in GLOBAL_ROLES:
        if str(prog["department_id"]) != user.get("department_id"):
            raise HTTPException(status_code=403, detail="Unauthorized for this department")

    # Credit cap per semester
    semester_subjects = await db.subjects.find({"program_id": subj.program_id, "semester": subj.semester}).to_list(100)
    current_credits = sum(s.get("credits", 0) for s in semester_subjects)
    if current_credits + subj.credits > 30:
        raise HTTPException(status_code=400, detail=f"Credit limit exceeded for Semester {subj.semester} (Max: 30, Current: {current_credits})")

    doc = {
        "name": subj.name,
        "code": subj.code,
        "type": subj.type,
        "credits": subj.credits,
        "program_id": subj.program_id,
        "semester": subj.semester,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.subjects.insert_one(doc)

    return SubjectResponse(
        id=str(result.inserted_id),
        name=subj.name,
        code=subj.code,
        type=subj.type,
        credits=subj.credits,
        program_id=subj.program_id,
        program_name=prog["name"],
        semester=subj.semester,
        created_at=doc["created_at"]
    )

@api_router.get("/subjects", response_model=List[SubjectResponse])
async def get_subjects(
    program_id: Optional[str] = None,
    semester: Optional[int] = None,
    type: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if program_id:
        query["program_id"] = program_id
    if semester:
        query["semester"] = semester
    if type:
        query["type"] = type
    
    # Department Isolation
    if user["role"] not in GLOBAL_ROLES and user["role"] != "student":
        if user.get("department_id"):
            # This requires filtering subjects by their program's department
            # For simplicity, we find all program IDs for this department first
            depts_progs = await db.programs.find({"department_id": user["department_id"]}).to_list(100)
            prog_ids = [str(p["_id"]) for p in depts_progs]
            if program_id and program_id not in prog_ids:
                return [] # Requested program outside permission
            query["program_id"] = {"$in": prog_ids}

    subjects = await db.subjects.find(query).to_list(1000)
    result = []
    for s in subjects:
        prog = await db.programs.find_one({"_id": ObjectId(s["program_id"]) if ObjectId.is_valid(s["program_id"]) else s["program_id"]})
        result.append(SubjectResponse(
            id=str(s["_id"]),
            name=s["name"],
            code=s["code"],
            credits=s.get("credits", 0),
            type=s.get("type", "Core"),
            program_id=s["program_id"],
            program_name=prog["name"] if prog else "Unknown",
            semester=s["semester"],
            created_at=s["created_at"]
        ))
    return result

@api_router.post("/electives/select")
async def select_elective(selection: ElectiveSelection, user: dict = Depends(require_roles(["student"]))):
    # Verify subject is an Elective for their program/semester
    student = await db.students.find_one({"email": user["email"]})
    subj = await db.subjects.find_one({"_id": ObjectId(selection.subject_id), "type": "Elective"})
    
    if not subj or subj["program_id"] != str(student["program_id"]):
        raise HTTPException(status_code=400, detail="Invalid elective selection")
        
    doc = {
        "student_id": str(student["_id"]),
        "subject_id": selection.subject_id,
        "semester": selection.semester,
        "status": "approved", # Auto-approve for now
        "selected_at": datetime.now(timezone.utc)
    }
    await db.elective_selections.insert_one(doc)
    return {"message": "Elective selected successfully"}

@api_router.delete("/subjects/{subj_id}")
async def delete_subject(subj_id: str, user: dict = Depends(require_roles(["admin", "dean", "dean_academics"]))):
    result = await db.subjects.delete_one({"_id": ObjectId(subj_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Subject not found")
    return {"message": "Subject deleted"}

# ==================== TIMETABLE ENDPOINTS ====================
@api_router.post("/timetable", response_model=TimetableSlotResponse)
async def create_timetable_slot(slot: TimetableSlotCreate, user: dict = Depends(require_roles(["admin", "dean", "dean_academics"]))):
    # Check for conflicts
    conflict = await db.timetable.find_one({
        "day": slot.day,
        "program_id": slot.program_id,
        "semester": slot.semester,
        "$or": [
            {"start_time": {"$lt": slot.end_time}, "end_time": {"$gt": slot.start_time}},
        ]
    })
    if conflict:
        raise HTTPException(status_code=400, detail="Time slot conflict detected")
    
    # Check faculty conflict
    faculty_conflict = await db.timetable.find_one({
        "day": slot.day,
        "faculty_id": slot.faculty_id,
        "$or": [
            {"start_time": {"$lt": slot.end_time}, "end_time": {"$gt": slot.start_time}},
        ]
    })
    if faculty_conflict:
        raise HTTPException(status_code=400, detail="Faculty already assigned at this time")
    
    # Check room conflict
    room_conflict = await db.timetable.find_one({
        "day": slot.day,
        "room": slot.room,
        "$or": [
            {"start_time": {"$lt": slot.end_time}, "end_time": {"$gt": slot.start_time}},
        ]
    })
    if room_conflict:
        raise HTTPException(status_code=400, detail="Room already booked at this time")
    
    subj = await db.subjects.find_one({"_id": ObjectId(slot.subject_id)})
    faculty = await db.users.find_one({"_id": ObjectId(slot.faculty_id)})
    prog = await db.programs.find_one({"_id": ObjectId(slot.program_id)})
    
    doc = {
        "day": slot.day,
        "start_time": slot.start_time,
        "end_time": slot.end_time,
        "subject_id": slot.subject_id,
        "faculty_id": slot.faculty_id,
        "room": slot.room,
        "program_id": slot.program_id,
        "semester": slot.semester,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.timetable.insert_one(doc)
    
    return TimetableSlotResponse(
        id=str(result.inserted_id),
        day=slot.day,
        start_time=slot.start_time,
        end_time=slot.end_time,
        subject_id=slot.subject_id,
        subject_name=subj["name"] if subj else None,
        faculty_id=slot.faculty_id,
        faculty_name=faculty["name"] if faculty else None,
        room=slot.room,
        program_id=slot.program_id,
        program_name=prog["name"] if prog else None,
        semester=slot.semester,
        created_at=doc["created_at"]
    )

@api_router.post("/timetable/bulk-import")
async def bulk_import_timetable(
    file: UploadFile = File(...),
    user: dict = Depends(require_roles(["admin", "dean", "dean_academics"]))
):
    content = await file.read()
    rows = []
    
    if file.filename.endswith(".xlsx"):
        try:
            wb = load_workbook(io.BytesIO(content))
            sheet = wb.active
            headers = [str(cell.value).lower().strip() for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):
                    row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                    rows.append(row_dict)
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading Excel: {str(e)}")
    else:
        try:
            decoded = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = [{k.lower().strip(): v for k, v in row.items()} for row in reader]
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading CSV: {str(e)}")

    imported = 0
    errors = []
    
    for idx, row in enumerate(rows):
        try:
            prog_code = str(row.get("program_code", "")).strip().upper()
            subj_code = str(row.get("subject_code", "")).strip().upper()
            email = str(row.get("faculty_email", "")).strip().lower()
            day = str(row.get("day", "")).strip().capitalize()
            start = str(row.get("start_time", "")).strip()
            end = str(row.get("end_time", "")).strip()
            room = str(row.get("room", "")).strip()
            sem = int(row.get("semester", 1))

            prog = await db.programs.find_one({"code": prog_code})
            subj = await db.subjects.find_one({"code": subj_code})
            faculty = await db.users.find_one({"email": email})

            if not prog: errors.append(f"Row {idx+2}: Program {prog_code} not found")
            elif not subj: errors.append(f"Row {idx+2}: Subject {subj_code} not found")
            elif not faculty: errors.append(f"Row {idx+2}: Faculty {email} not found")
            else:
                # Check for conflicts
                conflict = await db.timetable.find_one({
                    "day": day, "room": room,
                    "$or": [{"start_time": {"$lt": end}, "end_time": {"$gt": start}}]
                })
                if conflict:
                    errors.append(f"Row {idx+2}: Room conflict for {room} at {start}-{end}")
                    continue

                doc = {
                    "day": day, "start_time": start, "end_time": end,
                    "subject_id": str(subj["_id"]), "faculty_id": str(faculty["_id"]),
                    "room": room, "program_id": str(prog["_id"]), "semester": sem,
                    "created_at": datetime.now(timezone.utc)
                }
                await db.timetable.insert_one(doc)
                imported += 1
        except Exception as e:
            errors.append(f"Row {idx+2}: {str(e)}")

    return {"message": f"Successfully imported {imported} slots", "errors": errors}

@api_router.get("/timetable", response_model=List[TimetableSlotResponse])
async def get_timetable(
    program_id: Optional[str] = None,
    semester: Optional[int] = None,
    faculty_id: Optional[str] = None,
    day: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if program_id:
        query["program_id"] = program_id
    if semester:
        query["semester"] = semester
    if faculty_id:
        query["faculty_id"] = faculty_id
    if day:
        query["day"] = day
    
    slots = await db.timetable.find(query).to_list(500)
    result = []
    for slot in slots:
        subj = await db.subjects.find_one({"_id": ObjectId(slot["subject_id"])}) if slot.get("subject_id") else None
        faculty = await db.users.find_one({"_id": ObjectId(slot["faculty_id"])}) if slot.get("faculty_id") else None
        prog = await db.programs.find_one({"_id": ObjectId(slot["program_id"])}) if slot.get("program_id") else None
        result.append(TimetableSlotResponse(
            id=str(slot["_id"]),
            day=slot["day"],
            start_time=slot["start_time"],
            end_time=slot["end_time"],
            subject_id=slot["subject_id"],
            subject_name=subj["name"] if subj else None,
            faculty_id=slot["faculty_id"],
            faculty_name=faculty["name"] if faculty else None,
            room=slot["room"],
            program_id=slot["program_id"],
            program_name=prog["name"] if prog else None,
            semester=slot["semester"],
            created_at=slot["created_at"]
        ))
    return result

@api_router.delete("/timetable/{slot_id}")
async def delete_timetable_slot(slot_id: str, user: dict = Depends(require_roles(["admin", "dean", "dean_academics"]))):
    result = await db.timetable.delete_one({"_id": ObjectId(slot_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Slot not found")
    return {"message": "Timetable slot deleted"}

# ==================== ATTENDANCE ENDPOINTS ====================
@api_router.post("/attendance", response_model=AttendanceResponse)
async def mark_attendance(att: AttendanceCreate, user: dict = Depends(require_roles(["admin", "dean", "faculty", "staff"]))):
    student = await db.students.find_one({"_id": ObjectId(att.student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    subj = await db.subjects.find_one({"_id": ObjectId(att.subject_id)})
    if not subj:
        raise HTTPException(status_code=404, detail="Subject not found")
    
    # Check if already marked
    existing = await db.attendance.find_one({
        "student_id": att.student_id,
        "subject_id": att.subject_id,
        "date": att.date
    })
    
    if existing:
        # Update existing
        await db.attendance.update_one(
            {"_id": existing["_id"]},
            {"$set": {"status": att.status.value, "marked_by": user["id"]}}
        )
        return AttendanceResponse(
            id=str(existing["_id"]),
            student_id=att.student_id,
            student_name=student["name"],
            subject_id=att.subject_id,
            subject_name=subj["name"],
            date=att.date,
            status=att.status.value,
            marked_by=user["id"],
            created_at=existing["created_at"]
        )
    
    doc = {
        "student_id": att.student_id,
        "subject_id": att.subject_id,
        "date": att.date,
        "status": att.status.value,
        "marked_by": user["id"],
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.attendance.insert_one(doc)
    
    return AttendanceResponse(
        id=str(result.inserted_id),
        student_id=att.student_id,
        student_name=student["name"],
        subject_id=att.subject_id,
        subject_name=subj["name"],
        date=att.date,
        status=att.status.value,
        marked_by=user["id"],
        created_at=doc["created_at"]
    )

@api_router.post("/attendance/bulk")
async def bulk_mark_attendance(data: AttendanceBulkCreate, user: dict = Depends(require_roles(["admin", "dean", "faculty", "staff"]))):
    marked = 0
    for record in data.records:
        student_id = record.get("student_id")
        status = record.get("status", "present")
        
        existing = await db.attendance.find_one({
            "student_id": student_id,
            "subject_id": data.subject_id,
            "date": data.date
        })
        
        if existing:
            await db.attendance.update_one(
                {"_id": existing["_id"]},
                {"$set": {"status": status, "marked_by": user["id"]}}
            )
        else:
            await db.attendance.insert_one({
                "student_id": student_id,
                "subject_id": data.subject_id,
                "date": data.date,
                "status": status,
                "marked_by": user["id"],
                "created_at": datetime.now(timezone.utc)
            })
        marked += 1
    
    return {"marked": marked}

@api_router.get("/attendance", response_model=List[AttendanceResponse])
async def get_attendance(
    student_id: Optional[str] = None,
    subject_id: Optional[str] = None,
    date: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if student_id:
        query["student_id"] = student_id
    if subject_id:
        query["subject_id"] = subject_id
    if date:
        query["date"] = date
    
    records = await db.attendance.find(query).to_list(5000)
    result = []
    for r in records:
        student = await db.students.find_one({"_id": ObjectId(r["student_id"])}) if r.get("student_id") else None
        subj = await db.subjects.find_one({"_id": ObjectId(r["subject_id"])}) if r.get("subject_id") else None
        result.append(AttendanceResponse(
            id=str(r["_id"]),
            student_id=r["student_id"],
            student_name=student["name"] if student else None,
            subject_id=r["subject_id"],
            subject_name=subj["name"] if subj else None,
            date=r["date"],
            status=r["status"],
            marked_by=r.get("marked_by"),
            created_at=r["created_at"]
        ))
    return result

@api_router.get("/attendance/defaulters")
async def get_defaulters(
    threshold: float = 75.0,
    department_id: Optional[str] = None,
    program_id: Optional[str] = None,
    user: dict = Depends(require_roles(["admin", "dean", "dean_academics"]))
):
    query = {}
    if user["role"] == "dean":
        query["department_id"] = user["department_id"]
    elif department_id:
        query["department_id"] = department_id
        
    if program_id:
        query["program_id"] = program_id
        
    students = await db.students.find(query).to_list(1000)
    defaulters = []
    
    for s in students:
        records = await db.attendance.find({"student_id": str(s["_id"])}).to_list(5000)
        total = len(records)
        present = len([r for r in records if r["status"] in ["present", "late", "excused"]])
        
        perc = (present / total * 100) if total > 0 else 100 # Assume 100 if no classes yet? or 0? 100 is safer for defaulters
        if total > 0 and perc < threshold:
            prog = await db.programs.find_one({"_id": ObjectId(s["program_id"])})
            defaulters.append({
                "student_id": s["student_id"],
                "name": s["name"],
                "program": prog["name"] if prog else "N/A",
                "percentage": round(perc, 2),
                "total_conducted": total,
                "total_attended": present
            })
            
    return defaulters

@api_router.get("/attendance/stats/{student_id}")
async def get_attendance_stats(student_id: str, user: dict = Depends(get_current_user)):
    # Student Data Isolation
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
        
    if user["role"] == "student" and user["email"].lower() != student["email"].lower():
        raise HTTPException(status_code=403, detail="Unauthorized")

    records = await db.attendance.find({"student_id": student_id}).to_list(5000)
    
    # Group by subject
    stats = {}
    for r in records:
        subj_id = r["subject_id"]
        if subj_id not in stats:
            subj = await db.subjects.find_one({"_id": ObjectId(subj_id)})
            stats[subj_id] = {
                "subject_name": subj["name"] if subj else "Unknown",
                "subject_code": subj["code"] if subj else "N/A",
                "conducted": 0,
                "attended": 0,
                "absent": 0,
                "late": 0,
                "excused": 0
            }
        stats[subj_id]["conducted"] += 1
        if r["status"] in ["present", "late", "excused"]:
            stats[subj_id]["attended"] += 1
        
        if r["status"] == "absent": stats[subj_id]["absent"] += 1
        elif r["status"] == "late": stats[subj_id]["late"] += 1
        elif r["status"] == "excused": stats[subj_id]["excused"] += 1
    
    # Calculate percentages
    result = []
    for subj_id, data in stats.items():
        percentage = (data["attended"] / data["conducted"] * 100) if data["conducted"] > 0 else 0
        result.append({
            "subject_id": subj_id,
            "subject_name": data["subject_name"],
            "subject_code": data["subject_code"],
            "total_conducted": data["conducted"],
            "total_attended": data["attended"],
            "absent": data["absent"],
            "late": data["late"],
            "excused": data["excused"],
            "percentage": round(percentage, 2),
            "is_defaulter": percentage < 75.0
        })
    
    return result

@api_router.post("/attendance/staff/check-in")
async def staff_check_in(user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    now = datetime.now(timezone.utc)
    
    existing = await db.staff_attendance.find_one({"user_id": user["id"], "date": today})
    if existing:
        return {"message": "Already checked in", "check_in": existing["check_in"]}
        
    doc = {
        "user_id": user["id"],
        "user_name": user["name"],
        "role": user["role"],
        "department_id": user.get("department_id"),
        "date": today,
        "check_in": now,
        "check_out": None,
        "total_hours": 0,
        "status": "present",
        "created_at": now
    }
    await db.staff_attendance.insert_one(doc)
    return {"message": "Check-in successful", "check_in": now}

@api_router.post("/attendance/staff/check-out")
async def staff_check_out(user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    now = datetime.now(timezone.utc)
    
    existing = await db.staff_attendance.find_one({"user_id": user["id"], "date": today})
    if not existing:
        raise HTTPException(status_code=400, detail="No check-in found for today")
        
    # Calculate hours
    delta = now - existing["check_in"].replace(tzinfo=timezone.utc)
    hours = round(delta.total_seconds() / 3600, 2)
    
    await db.staff_attendance.update_one(
        {"_id": existing["_id"]},
        {"$set": {"check_out": now, "total_hours": hours}}
    )
    return {"message": "Check-out successful", "total_hours": hours}

@api_router.get("/attendance/staff")
async def get_staff_attendance(
    department_id: Optional[str] = None,
    role: Optional[str] = None,
    date: Optional[str] = None,
    user: dict = Depends(require_roles(["admin", "dean", "dean_academics"]))
):
    query = {}
    if user["role"] == "dean":
        query["department_id"] = user["department_id"]
    elif department_id:
        query["department_id"] = department_id
        
    if role: query["role"] = role
    if date: query["date"] = date
    
    records = await db.staff_attendance.find(query).to_list(1000)
    for r in records: r["id"] = str(r.pop("_id"))
    return records
@api_router.post("/electives", response_model=ElectiveResponse)
async def create_elective(elec: ElectiveCreate, user: dict = Depends(require_roles(["admin", "dean", "dean_academics"]))):
    existing = await db.electives.find_one({"code": elec.code})
    if existing:
        raise HTTPException(status_code=400, detail="Elective code already exists")
    
    prog = await db.programs.find_one({"_id": ObjectId(elec.program_id)})
    if not prog:
        raise HTTPException(status_code=400, detail="Program not found")
    
    doc = {
        "name": elec.name,
        "code": elec.code,
        "credits": elec.credits,
        "program_id": elec.program_id,
        "semester": elec.semester,
        "max_students": elec.max_students,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.electives.insert_one(doc)
    
    return ElectiveResponse(
        id=str(result.inserted_id),
        name=elec.name,
        code=elec.code,
        credits=elec.credits,
        program_id=elec.program_id,
        program_name=prog["name"],
        semester=elec.semester,
        max_students=elec.max_students,
        created_at=doc["created_at"]
    )

@api_router.get("/electives", response_model=List[ElectiveResponse])
async def get_electives(
    program_id: Optional[str] = None,
    semester: Optional[int] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if program_id:
        query["program_id"] = program_id
    if semester:
        query["semester"] = semester
    
    electives = await db.electives.find(query).to_list(200)
    result = []
    for e in electives:
        prog = await db.programs.find_one({"_id": ObjectId(e["program_id"])}) if e.get("program_id") else None
        result.append(ElectiveResponse(
            id=str(e["_id"]),
            name=e["name"],
            code=e["code"],
            credits=e["credits"],
            program_id=e["program_id"],
            program_name=prog["name"] if prog else None,
            semester=e["semester"],
            max_students=e["max_students"],
            created_at=e["created_at"]
        ))
    return result

# ==================== BATCH ENDPOINTS ====================
@api_router.post("/batches", response_model=BatchResponse)
async def create_batch(batch: BatchCreate, user: dict = Depends(require_roles(["admin", "dean", "dean_academics"]))):
    elec = await db.electives.find_one({"_id": ObjectId(batch.elective_id)})
    if not elec:
        raise HTTPException(status_code=400, detail="Elective not found")
    
    faculty = await db.users.find_one({"_id": ObjectId(batch.faculty_id)})
    if not faculty:
        raise HTTPException(status_code=400, detail="Faculty not found")
    
    doc = {
        "elective_id": batch.elective_id,
        "faculty_id": batch.faculty_id,
        "batch_name": batch.batch_name,
        "student_ids": [],
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.batches.insert_one(doc)
    
    return BatchResponse(
        id=str(result.inserted_id),
        elective_id=batch.elective_id,
        elective_name=elec["name"],
        faculty_id=batch.faculty_id,
        faculty_name=faculty["name"],
        batch_name=batch.batch_name,
        student_ids=[],
        created_at=doc["created_at"]
    )

@api_router.post("/batches/{batch_id}/students/{student_id}")
async def add_student_to_batch(batch_id: str, student_id: str, user: dict = Depends(require_roles(["admin", "dean", "dean_academics", "staff"]))):
    batch = await db.batches.find_one({"_id": ObjectId(batch_id)})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    if student_id in batch.get("student_ids", []):
        raise HTTPException(status_code=400, detail="Student already in batch")
    
    await db.batches.update_one(
        {"_id": ObjectId(batch_id)},
        {"$push": {"student_ids": student_id}}
    )
    return {"message": "Student added to batch"}

@api_router.delete("/batches/{batch_id}/students/{student_id}")
async def remove_student_from_batch(batch_id: str, student_id: str, user: dict = Depends(require_roles(["admin", "dean", "dean_academics", "staff"]))):
    await db.batches.update_one(
        {"_id": ObjectId(batch_id)},
        {"$pull": {"student_ids": student_id}}
    )
    return {"message": "Student removed from batch"}

@api_router.get("/batches", response_model=List[BatchResponse])
async def get_batches(
    elective_id: Optional[str] = None,
    faculty_id: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if elective_id:
        query["elective_id"] = elective_id
    if faculty_id:
        query["faculty_id"] = faculty_id
    
    batches = await db.batches.find(query).to_list(200)
    result = []
    for b in batches:
        elec = await db.electives.find_one({"_id": ObjectId(b["elective_id"])}) if b.get("elective_id") else None
        faculty = await db.users.find_one({"_id": ObjectId(b["faculty_id"])}) if b.get("faculty_id") else None
        result.append(BatchResponse(
            id=str(b["_id"]),
            elective_id=b["elective_id"],
            elective_name=elec["name"] if elec else None,
            faculty_id=b["faculty_id"],
            faculty_name=faculty["name"] if faculty else None,
            batch_name=b["batch_name"],
            student_ids=b.get("student_ids", []),
            created_at=b["created_at"]
        ))
    return result

# ==================== USER MANAGEMENT ====================
@api_router.get("/users", response_model=List[UserResponse])
async def get_users(
    role: Optional[str] = None,
    department_id: Optional[str] = None,
    user: dict = Depends(require_roles(["admin", "dean"]))
):
    query = {}
    if role:
        query["role"] = role
    if department_id:
        query["department_id"] = department_id
    
    users = await db.users.find(query, {"password_hash": 0}).to_list(500)
    return [UserResponse(
        id=str(u["_id"]),
        email=u["email"],
        name=u["name"],
        role=u["role"],
        department_id=u.get("department_id"),
        created_at=u["created_at"]
    ) for u in users]

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, role: str, user: dict = Depends(require_roles(["admin"]))):
    if role not in [r.value for r in UserRole]:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    result = await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"role": role}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "Role updated"}

# ==================== DASHBOARD STATS ====================
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user: dict = Depends(get_current_user)):
    query = {}
    if user["role"] not in GLOBAL_ROLES and user["role"] != "student":
        if user.get("department_id"):
            query["department_id"] = user["department_id"]

    faculty_query = {"role": "faculty"}
    if user["role"] not in GLOBAL_ROLES and user["role"] != "student" and user.get("department_id"):
        faculty_query["department_id"] = user["department_id"]

    total_students, total_departments, total_programs, total_faculty, total_subjects = await asyncio.gather(
        db.students.count_documents(query),
        db.departments.count_documents(query),
        db.programs.count_documents(query),
        db.users.count_documents(faculty_query),
        db.subjects.count_documents({}),
    )

    return {
        "total_students": total_students,
        "total_faculty": total_faculty,
        "total_departments": total_departments,
        "total_programs": total_programs,
        "total_subjects": total_subjects
    }

@api_router.get("/dashboard/hierarchy")
async def get_dashboard_hierarchy(user: dict = Depends(get_current_user)):
    """
    Returns departments with their programs and student counts.
    Uses a single $group aggregation — much faster than fetching all students.
    """
    # 1. Count students per program_id in one aggregation query
    pipeline = [{"$group": {"_id": "$program_id", "count": {"$sum": 1}}}]
    counts_cursor = db.students.aggregate(pipeline)
    student_counts = {doc["_id"]: doc["count"] async for doc in counts_cursor}

    # 2. Fetch departments & programs in parallel (2 queries total)
    depts_raw, progs_raw = await asyncio.gather(
        db.departments.find({}).to_list(100),
        db.programs.find({}).to_list(500),
    )

    # 3. Group programs by department
    programs_by_dept = {}
    for p in progs_raw:
        did = str(p.get("department_id", ""))
        programs_by_dept.setdefault(did, []).append({
            "id": str(p["_id"]),
            "name": p["name"],
            "code": p["code"],
            "student_count": student_counts.get(str(p["_id"]), 0)
        })

    # 4. Build response
    result = []
    for d in depts_raw:
        did = str(d["_id"])
        result.append({
            "id": did,
            "name": d["name"],
            "code": d["code"],
            "programs": programs_by_dept.get(did, [])
        })
    return result

# ==================== STARTUP ====================
@app.on_event("startup")
async def startup():
    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.departments.create_index("code", unique=True)
    await db.programs.create_index("code", unique=True)
    await db.students.create_index("enrollment_number", unique=True)
    await db.students.create_index("student_id", unique=True)
    await db.subjects.create_index("code", unique=True)
    await db.electives.create_index("code", unique=True)
    
    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@raffles.edu.in")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "System Administrator",
            "role": "admin",
            "department_id": None,
            "created_at": datetime.now(timezone.utc)
        })
        logger.info(f"Admin user created: {admin_email}")
    
    # Seed Departments
    departments = [
        {"name": "SCHOOL OF LAW", "code": "LAW"},
        {"name": "SCHOOL OF ENGINEERING & TECHNOLOGY", "code": "SET"},
        {"name": "SCHOOL OF BASIC & APPLIED SCIENCES", "code": "BAS"},
        {"name": "SCHOOL OF AGRICULTURAL SCIENCES", "code": "AGR"},
        {"name": "ALABBAR SCHOOL OF MANAGEMENT", "code": "ASM"},
        {"name": "SCHOOL OF PHARMACY", "code": "PHM"},
        {"name": "SCHOOL OF HUMANITIES & SOCIAL SCIENCES", "code": "HSS"}
    ]
    
    for dept in departments:
        existing_dept = await db.departments.find_one({"code": dept["code"]})
        if not existing_dept:
            await db.departments.insert_one({
                "name": dept["name"],
                "code": dept["code"],
                "head_id": None,
                "created_at": datetime.now(timezone.utc)
            })
            logger.info(f"Department created: {dept['name']}")
    
    # Seed Programs
    programs = [
        {"name": "BA LLB", "code": "BALLB", "dept": "LAW", "duration": 5, "sem": 10},
        {"name": "BBA LLB", "code": "BBALLB", "dept": "LAW", "duration": 5, "sem": 10},
        {"name": "LLB.", "code": "LLB", "dept": "LAW", "duration": 3, "sem": 6},
        {"name": "LLM", "code": "LLM", "dept": "LAW", "duration": 2, "sem": 4},
        {"name": "Certificate in Legal Studies", "code": "CLS", "dept": "LAW", "duration": 1, "sem": 2},
        
        {"name": "DIPLOMA", "code": "DIP", "dept": "SET", "duration": 3, "sem": 6},
        {"name": "BCA", "code": "BCA", "dept": "SET", "duration": 3, "sem": 6},
        {"name": "B.TECH", "code": "BTECH", "dept": "SET", "duration": 4, "sem": 8},
        {"name": "MCA", "code": "MCA", "dept": "SET", "duration": 2, "sem": 4},
        {"name": "M.TECH", "code": "MTECH", "dept": "SET", "duration": 2, "sem": 4},
        {"name": "PG Diploma in AI", "code": "PGDAI", "dept": "SET", "duration": 1, "sem": 2},
        
        {"name": "B.SC", "code": "BSC", "dept": "BAS", "duration": 3, "sem": 6},
        {"name": "M.SC", "code": "MSC", "dept": "BAS", "duration": 2, "sem": 4},
        
        {"name": "B.SC (AGRICULTURE)", "code": "BSCAGRI", "dept": "AGR", "duration": 4, "sem": 8},
        
        {"name": "B.COM", "code": "BCOM", "dept": "ASM", "duration": 3, "sem": 6},
        {"name": "BBA.", "code": "BBA", "dept": "ASM", "duration": 3, "sem": 6},
        {"name": "M.COM", "code": "MCOM", "dept": "ASM", "duration": 2, "sem": 4},
        {"name": "MBA", "code": "MBA", "dept": "ASM", "duration": 2, "sem": 4},
        {"name": "Ph.D in Management", "code": "PHDMGT", "dept": "ASM", "duration": 3, "sem": 6},
        
        {"name": "D.PHARM", "code": "DPHARM", "dept": "PHM", "duration": 2, "sem": 4},
        {"name": "B.PHARM", "code": "BPHARM", "dept": "PHM", "duration": 4, "sem": 8},
        {"name": "M.PHARM", "code": "MPHARM", "dept": "PHM", "duration": 2, "sem": 4},
        
        {"name": "B.A (HONS)", "code": "BAHONS", "dept": "HSS", "duration": 3, "sem": 6},
        {"name": "B.A (GEN)", "code": "BAGEN", "dept": "HSS", "duration": 3, "sem": 6},
        {"name": "M.A.", "code": "MA", "dept": "HSS", "duration": 2, "sem": 4}
    ]
    
    for prog in programs:
        existing_prog = await db.programs.find_one({"code": prog["code"]})
        if not existing_prog:
            dept = await db.departments.find_one({"code": prog["dept"]})
            if dept:
                await db.programs.insert_one({
                    "name": prog["name"],
                    "code": prog["code"],
                    "department_id": str(dept["_id"]),
                    "duration_years": prog["duration"],
                    "total_semesters": prog["sem"],
                    "created_at": datetime.now(timezone.utc)
                })
                logger.info(f"Program created: {prog['name']}")
    
    # Seed Subjects for B.Tech (SET)
    btech = await db.programs.find_one({"code": "BTECH"})
    if btech:
        subjects_to_seed = [
            {"name": "Programming in C", "code": "CS101", "type": "Core", "credits": 4.0, "sem": 1},
            {"name": "Mathematics I", "code": "MA101", "type": "Core", "credits": 3.0, "sem": 1},
            {"name": "Engineering Graphics", "code": "ME101", "type": "Lab", "credits": 2.0, "sem": 1},
            {"name": "Data Structures", "code": "CS201", "type": "Core", "credits": 4.0, "sem": 3},
            {"name": "Introduction to AI", "code": "CS205", "type": "Elective", "credits": 3.0, "sem": 3},
        ]
        for s in subjects_to_seed:
            await db.subjects.update_one(
                {"code": s["code"]},
                {"$set": {
                    "name": s["name"],
                    "type": s["type"],
                    "credits": s["credits"],
                    "program_id": str(btech["_id"]),
                    "semester": s["sem"],
                    "created_at": datetime.now(timezone.utc)
                }},
                upsert=True
            )
            logger.info(f"Subject Upserted: {s['name']}")
    
    # Write test credentials
    memo_path = Path(__file__).parent / "memory"
    memo_path.mkdir(exist_ok=True)
    with open(memo_path / "test_credentials.md", "w") as f:
        f.write("# Test Credentials\n\n")
        f.write(f"## Admin Account\n")
        f.write(f"- Email: {admin_email}\n")
        f.write(f"- Password: {admin_password}\n")
        f.write(f"- Role: admin\n\n")
        f.write("## Auth Endpoints\n")
        f.write("- POST /api/auth/login\n")
        f.write("- POST /api/auth/register\n")
        f.write("- POST /api/auth/logout\n")
        f.write("- GET /api/auth/me\n")

# Include router
app.include_router(api_router, prefix="/api")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[os.environ.get('FRONTEND_URL', 'http://localhost:3000'), os.environ.get('CORS_ORIGINS', '*')],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# Global exception handler to expose traceback during debugging
import traceback as _traceback
from fastapi.responses import JSONResponse

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    tb = _traceback.format_exc()
    logger.error(f"Unhandled exception on {request.method} {request.url}:\n{tb}")
    return JSONResponse(
        status_code=500,
        content={"detail": str(exc), "traceback": tb}
    )
